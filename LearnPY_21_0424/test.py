import unittest
from HTMLTestRunner import HTMLTestRunner

import openpyxl
from ddt import ddt, data, unpack
from selenium import webdriver
import os
import time


# 读取excel
# xlrd,xlwt,openpyxl
def read_excel():
    workbook = openpyxl.load_workbook("data.xlsx")
    sheet = workbook["login"]
    print(sheet.max_row, sheet.max_column)
    allList = []
    for row in range(2, sheet.max_row + 1):   # 行数
        templist = []
        for col in range(1, sheet.max_column + 1):  # 列数
            templist.append(sheet.cell(row, col).value)
        allList.append(templist)
    return allList


@ddt
class Test(unittest.TestCase):

    @data(*read_excel())     # 加星号表示以列表的形式去解析
    @unpack
    def test_jenkins_login(self, order, username, password):
        global driver
        driver = webdriver.Firefox()
        driver.get("http://localhost:8080/sds/public/index.html")
        driver.find_element_by_xpath("/html/body/div[1]/form/div[2]/div/input").send_keys(username)
        driver.find_element_by_xpath("/html/body/div[1]/form/div[3]/div/input").send_keys(password)
        driver.find_element_by_xpath("/html/body/div[1]/form/div[6]/input[2]").click()
        # 断言
        if order == 1:
            self.assertEqual(1, 1)
        elif order == 2:
            self.assertEqual(1, 1)
        else:
            self.assertEqual(1, 1)


if __name__ == '__main__':
    testcases = unittest.defaultTestLoader.discover(os.getcwd(), "*.py")
    filename = open(os.getcwd()+"/report.html", "wb")
    runner = HTMLTestRunner(stream=filename, verbosity=2,
                            title="2021/04/24测试报告", description="报告详情如下：")
    runner.run(testcases)